"""Smoke + dry-run tests for scripts/svf_linz_regression.py.

The full regression is too slow for casual CI; these tests use a
30-participant synthetic CSV and the script's own --dry-run mode.
"""

from __future__ import annotations

import subprocess
import sys
import textwrap
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

REPO_ROOT = Path(__file__).resolve().parents[1]
SCRIPT = REPO_ROOT / "scripts" / "svf_linz_regression.py"


KELLY_FIXTURE_XML = """<?xml version="1.0" encoding="UTF-8"?>
<LexicalResource>
  <Lexicon>
    <LexicalEntry><Lemma><FormRepresentation>
      <feat att="writtenForm" val="hund"/>
      <feat att="kellyPartOfSpeech" val="noun-en"/>
      <feat att="rawFreq" val="500"/><feat att="wpm" val="500,00"/>
      <feat att="cefr" val="1"/><feat att="source" val="corpus"/>
    </FormRepresentation></Lemma></LexicalEntry>
    <LexicalEntry><Lemma><FormRepresentation>
      <feat att="writtenForm" val="katt"/>
      <feat att="kellyPartOfSpeech" val="noun-en"/>
      <feat att="rawFreq" val="300"/><feat att="wpm" val="300,00"/>
      <feat att="cefr" val="1"/><feat att="source" val="corpus"/>
    </FormRepresentation></Lemma></LexicalEntry>
    <LexicalEntry><Lemma><FormRepresentation>
      <feat att="writtenForm" val="häst"/>
      <feat att="kellyPartOfSpeech" val="noun-en"/>
      <feat att="rawFreq" val="200"/><feat att="wpm" val="200,00"/>
      <feat att="cefr" val="2"/><feat att="source" val="corpus"/>
    </FormRepresentation></Lemma></LexicalEntry>
    <LexicalEntry><Lemma><FormRepresentation>
      <feat att="writtenForm" val="ko"/>
      <feat att="kellyPartOfSpeech" val="noun-en"/>
      <feat att="rawFreq" val="150"/><feat att="wpm" val="150,00"/>
      <feat att="cefr" val="2"/><feat att="source" val="corpus"/>
    </FormRepresentation></Lemma></LexicalEntry>
  </Lexicon>
</LexicalResource>
"""


def write_kelly_fixture(tmp_path: Path) -> Path:
    p = tmp_path / "kelly_fixture.xml"
    p.write_text(KELLY_FIXTURE_XML, encoding="utf-8")
    return p


def write_svf_xlsx(
    tmp_path: Path,
    participants: list[dict],
    filename: str = "svf_fixture.xlsx",
) -> Path:
    """Write a minimal SVF XLSX in the v3 layout the data loader expects.

    Columns: A=labels, B...= User-1, User-2, ...
    Row 0 (header): blank, User-1, User-2, ...
    Rows 1..R: responses (one per cell, None for blank)
    Then 'Gender:', 'Age:', 'Category:', 'MMSE' rows.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    user_ids = [p["participant_id"] for p in participants]
    ws.cell(row=1, column=1, value=None)
    for j, uid in enumerate(user_ids, start=2):
        ws.cell(row=1, column=j, value=uid)
    max_responses = max(len(p["responses"]) for p in participants)
    for i in range(max_responses):
        for j, p in enumerate(participants, start=2):
            if i < len(p["responses"]):
                ws.cell(row=2 + i, column=j, value=p["responses"][i])
    meta_row = 2 + max_responses
    ws.cell(row=meta_row, column=1, value="Gender:")
    for j, p in enumerate(participants, start=2):
        ws.cell(row=meta_row, column=j, value=p.get("gender", "F"))
    ws.cell(row=meta_row + 1, column=1, value="Age:")
    for j, p in enumerate(participants, start=2):
        ws.cell(row=meta_row + 1, column=j, value=float(p.get("age", 70.0)))
    ws.cell(row=meta_row + 2, column=1, value="Category:")
    for j, p in enumerate(participants, start=2):
        ws.cell(row=meta_row + 2, column=j, value=p["diagnosis"])
    ws.cell(row=meta_row + 3, column=1, value="MMSE")
    for j, p in enumerate(participants, start=2):
        ws.cell(row=meta_row + 3, column=j, value=float(p["mmse"]))
    out = tmp_path / filename
    wb.save(out)
    return out


def make_paired_fixture(tmp_path: Path) -> tuple[Path, Path, Path]:
    """Build a matching CSV + XLSX + Kelly XML so the recompute path runs.

    Returns (csv_path, xlsx_path, kelly_path). All four diagnoses get
    five participants. MMSE varies by group; total_words tracks MMSE so
    the |r|>0.3 threshold has at least one survivor regardless of which
    backend supplies MWF.
    """
    rng = np.random.default_rng(0)
    diagnoses = ["HC", "MCI", "non-AD", "AD"]
    diag_to_mmse_mean = {"HC": 28.5, "MCI": 25.5, "non-AD": 22.0, "AD": 18.0}
    word_pool = ["hund", "katt", "häst", "ko"]  # all in the Kelly fixture
    participants = []
    rows = []
    pid = 0
    for diag in diagnoses:
        for _ in range(5):
            pid += 1
            mmse = float(np.clip(
                rng.normal(diag_to_mmse_mean[diag], 1.5), 0, 30,
            ))
            wc = float(np.clip(rng.normal(mmse - 8, 2), 1, 30))
            n_resp = max(2, int(wc))
            responses = list(rng.choice(word_pool, size=n_resp, replace=True))
            unique_words = max(1, int(wc - rng.integers(0, 3)))
            participants.append({
                "participant_id": f"User-{pid}",
                "diagnosis": diag,
                "age": float(rng.integers(55, 85)),
                "gender": str(rng.choice(["M", "F"])),
                "mmse": mmse,
                "responses": responses,
            })
            rows.append({
                "participant_id": f"User-{pid}",
                "diagnosis": diag,
                "age": participants[-1]["age"],
                "gender": participants[-1]["gender"],
                "mmse": mmse,
                "total_words": wc,
                "unique_words": unique_words,
                "repetitions": int(wc) - unique_words,
                "mean_consecutive_similarity": float(rng.uniform(0.3, 0.7)),
                "pairwise_similarity_mean": float(rng.uniform(0.2, 0.5)),
                "temporal_gradient": float(rng.normal(0, 0.05)),
                "similarity_slope": float(rng.normal(0, 0.01)),
                "cluster_count": int(rng.integers(1, max(2, int(wc)))),
                "mean_cluster_size": float(rng.uniform(1.0, 3.0)),
                "switch_count": int(rng.integers(0, max(1, int(wc) - 1))),
                "max_cluster_size": int(rng.integers(1, max(2, int(wc)))),
                # Filler value; the recompute path overwrites this.
                "mean_word_frequency": 4.0,
            })
    csv_path = tmp_path / "svf_results_with_mmse.csv"
    pd.DataFrame(rows).to_csv(csv_path, index=False)
    xlsx_path = write_svf_xlsx(tmp_path, participants)
    kelly_path = write_kelly_fixture(tmp_path)
    return csv_path, xlsx_path, kelly_path


def make_synthetic_linz_csv(tmp_path: Path, n: int = 30) -> Path:
    """Build a small synthetic SVF results CSV with known-ish correlations.

    Keeps each diagnosis group large enough for 5-fold StratifiedKFold,
    so we balance ≥ 5 participants per diagnosis. WC and MMSE are
    correlated so the |r|>0.3 filter has at least one survivor.
    """
    rng = np.random.default_rng(0)
    diagnoses = ["HC", "MCI", "non-AD", "AD"]
    rows = []
    per_group = max(5, n // len(diagnoses))
    diag_to_mmse_mean = {"HC": 28.5, "MCI": 25.5, "non-AD": 22.0, "AD": 18.0}
    pid = 0
    for diag in diagnoses:
        for _ in range(per_group):
            pid += 1
            mmse_mean = diag_to_mmse_mean[diag]
            mmse = float(np.clip(rng.normal(mmse_mean, 1.5), 0, 30))
            wc = float(np.clip(rng.normal(mmse - 8, 2), 1, 30))
            unique_words = max(1, int(wc - rng.integers(0, 3)))
            rows.append({
                "participant_id": f"User-{pid}",
                "diagnosis": diag,
                "age": float(rng.integers(55, 85)),
                "gender": rng.choice(["M", "F"]),
                "mmse": mmse,
                "total_words": wc,
                "unique_words": unique_words,
                "repetitions": int(wc) - unique_words,
                "mean_consecutive_similarity": float(rng.uniform(0.3, 0.7)),
                "pairwise_similarity_mean": float(rng.uniform(0.2, 0.5)),
                "temporal_gradient": float(rng.normal(0, 0.05)),
                "similarity_slope": float(rng.normal(0, 0.01)),
                "cluster_count": int(rng.integers(1, max(2, int(wc)))),
                "mean_cluster_size": float(rng.uniform(1.0, 3.0)),
                "switch_count": int(rng.integers(0, max(1, int(wc) - 1))),
                "max_cluster_size": int(rng.integers(1, max(2, int(wc)))),
                "mean_word_frequency": float(np.clip(
                    rng.normal(4.0 + (mmse - 25) * 0.05, 0.3), 1.0, 6.0,
                )),
            })
    df = pd.DataFrame(rows)
    out = tmp_path / "svf_results_with_mmse.csv"
    df.to_csv(out, index=False)
    return out


def test_linz_regression_dry_run_lists_planned_analysis(tmp_path):
    """--dry-run prints the planned analysis without executing."""
    fixture = make_synthetic_linz_csv(tmp_path)
    result = subprocess.run(
        [
            sys.executable, str(SCRIPT),
            "--input", str(fixture),
            "--output-dir", str(tmp_path / "out"),
            "--dry-run",
        ],
        capture_output=True, text=True,
    )
    assert result.returncode == 0, result.stderr
    out = result.stdout
    assert "5 Linz features" in out
    assert "SVR" in out
    assert "Ridge" in out
    assert "1000 bootstrap" in out
    assert "MMSE = 24" in out


def test_linz_regression_end_to_end_smoke(tmp_path):
    """Full run on a small fixture; verify all four output artefacts exist."""
    fixture = make_synthetic_linz_csv(tmp_path)
    out_dir = tmp_path / "out"
    fig_dir = tmp_path / "fig"
    result = subprocess.run(
        [
            sys.executable, str(SCRIPT),
            "--input", str(fixture),
            "--output-dir", str(out_dir),
            "--figure-dir", str(fig_dir),
            "--random-seed", "42",
        ],
        capture_output=True, text=True,
    )
    assert result.returncode == 0, result.stderr
    assert (out_dir / "linz_correlation_table.csv").is_file()
    assert (out_dir / "linz_regression_mae.csv").is_file()
    assert (out_dir / "linz_confusion_matrix.csv").is_file()
    assert (out_dir / "linz_kappa.csv").is_file()
    assert (fig_dir / "svf_predicted_vs_actual.png").is_file()

    mae = pd.read_csv(out_dir / "linz_regression_mae.csv")
    assert set(mae["model"]) == {"SVR", "Ridge"}
    # Sanity: MAE on 0-30 scale should be well under the all-zero/30 floor.
    assert (mae["mae_mean"] < 8).all()


def test_linz_regression_errors_when_no_features_pass(tmp_path):
    """If every Linz feature is uncorrelated with MMSE, exit non-zero."""
    rng = np.random.default_rng(1)
    diagnoses = ["HC", "MCI", "non-AD", "AD"]
    rows = []
    pid = 0
    for diag in diagnoses:
        for _ in range(5):
            pid += 1
            rows.append({
                "participant_id": f"User-{pid}",
                "diagnosis": diag,
                "age": 70.0,
                "gender": "F",
                # mmse and features are independent random noise
                "mmse": float(rng.uniform(0, 30)),
                "total_words": float(rng.uniform(0, 30)),
                "unique_words": int(rng.integers(0, 30)),
                "repetitions": 0,
                "mean_consecutive_similarity": float(rng.uniform(0, 1)),
                "pairwise_similarity_mean": float(rng.uniform(0, 1)),
                "temporal_gradient": float(rng.normal(0, 1)),
                "similarity_slope": float(rng.normal(0, 1)),
                "cluster_count": int(rng.integers(1, 10)),
                "mean_cluster_size": float(rng.uniform(1, 3)),
                "switch_count": int(rng.integers(0, 10)),
                "max_cluster_size": int(rng.integers(1, 5)),
                "mean_word_frequency": float(rng.uniform(0, 6)),
            })
    df = pd.DataFrame(rows)
    fixture = tmp_path / "uncorrelated.csv"
    df.to_csv(fixture, index=False)

    result = subprocess.run(
        [
            sys.executable, str(SCRIPT),
            "--input", str(fixture),
            "--output-dir", str(tmp_path / "out"),
            "--figure-dir", str(tmp_path / "fig"),
            "--random-seed", "42",
        ],
        capture_output=True, text=True,
    )
    # Could pass (some random noise crosses 0.3) or fail with code 2 (the
    # designed exit). What we forbid is a silent crash on empty features.
    assert result.returncode in (0, 2), result.stderr


# ──────────────────────────────────────────────────────
# Kelly / wordfreq backend recompute paths
# ──────────────────────────────────────────────────────

def test_recompute_mwf_kelly_overwrites_csv_value(tmp_path):
    """Direct call to recompute_mwf with source='kelly' replaces the column."""
    sys.path.insert(0, str(REPO_ROOT))
    sys.path.insert(0, str(REPO_ROOT / "scripts"))
    import svf_linz_regression as mod  # noqa: E402

    csv_path, xlsx_path, kelly_path = make_paired_fixture(tmp_path)
    df = pd.read_csv(csv_path)
    original_mwf = df["mean_word_frequency"].copy()

    out = mod.recompute_mwf(df, xlsx_path, "kelly", kelly_path=kelly_path)
    assert "mean_word_frequency" in out.columns
    # Original CSV had the constant filler 4.0; recompute pulls real
    # Kelly Zipf scores from the fixture (~5.6–5.7 for hund/katt) so
    # at least one row must differ.
    assert (out["mean_word_frequency"] != original_mwf).any()
    assert out["mean_word_frequency"].notna().all()


def test_recompute_mwf_wordfreq_returns_finite(tmp_path):
    sys.path.insert(0, str(REPO_ROOT))
    sys.path.insert(0, str(REPO_ROOT / "scripts"))
    import svf_linz_regression as mod  # noqa: E402

    csv_path, xlsx_path, _ = make_paired_fixture(tmp_path)
    df = pd.read_csv(csv_path)
    out = mod.recompute_mwf(df, xlsx_path, "wordfreq")
    assert out["mean_word_frequency"].notna().all()
    assert (out["mean_word_frequency"] > 0).all()


def test_recompute_mwf_rejects_unknown_source(tmp_path):
    sys.path.insert(0, str(REPO_ROOT))
    sys.path.insert(0, str(REPO_ROOT / "scripts"))
    import svf_linz_regression as mod  # noqa: E402

    csv_path, xlsx_path, _ = make_paired_fixture(tmp_path)
    df = pd.read_csv(csv_path)
    import pytest as _pt
    with _pt.raises(ValueError, match="unsupported source"):
        mod.recompute_mwf(df, xlsx_path, "csv")


def test_recompute_mwf_raises_when_xlsx_missing(tmp_path):
    sys.path.insert(0, str(REPO_ROOT))
    sys.path.insert(0, str(REPO_ROOT / "scripts"))
    import svf_linz_regression as mod  # noqa: E402

    csv_path, _, _ = make_paired_fixture(tmp_path)
    df = pd.read_csv(csv_path)
    import pytest as _pt
    with _pt.raises(FileNotFoundError, match="--svf-data"):
        mod.recompute_mwf(df, tmp_path / "missing.xlsx", "kelly")


def test_recompute_mwf_raises_when_participant_missing(tmp_path):
    sys.path.insert(0, str(REPO_ROOT))
    sys.path.insert(0, str(REPO_ROOT / "scripts"))
    import svf_linz_regression as mod  # noqa: E402

    csv_path, xlsx_path, kelly_path = make_paired_fixture(tmp_path)
    df = pd.read_csv(csv_path)
    df.loc[len(df)] = df.iloc[0].copy()
    df.loc[len(df) - 1, "participant_id"] = "User-9999"
    import pytest as _pt
    with _pt.raises(ValueError, match="absent from"):
        mod.recompute_mwf(df, xlsx_path, "kelly", kelly_path=kelly_path)


def test_linz_regression_kelly_end_to_end(tmp_path):
    """End-to-end smoke for --frequency-source kelly."""
    csv_path, xlsx_path, kelly_path = make_paired_fixture(tmp_path)
    out_dir = tmp_path / "out"
    fig_dir = tmp_path / "fig"
    result = subprocess.run(
        [
            sys.executable, str(SCRIPT),
            "--input", str(csv_path),
            "--output-dir", str(out_dir),
            "--figure-dir", str(fig_dir),
            "--random-seed", "42",
            "--frequency-source", "kelly",
            "--svf-data", str(xlsx_path),
            "--kelly-path", str(kelly_path),
        ],
        capture_output=True, text=True,
    )
    # Either the regression runs (0) or it bails because no feature
    # passes the |r|>0.3 threshold (2). Both are valid for this random
    # fixture; the silent crash is what we forbid.
    assert result.returncode in (0, 2), result.stderr


def test_linz_regression_compare_frequency_sources(tmp_path):
    """--compare-frequency-sources writes per-source dirs and a comparison CSV."""
    csv_path, xlsx_path, kelly_path = make_paired_fixture(tmp_path)
    out_dir = tmp_path / "out"
    fig_dir = tmp_path / "fig"
    result = subprocess.run(
        [
            sys.executable, str(SCRIPT),
            "--input", str(csv_path),
            "--output-dir", str(out_dir),
            "--figure-dir", str(fig_dir),
            "--random-seed", "42",
            "--compare-frequency-sources",
            "--svf-data", str(xlsx_path),
            "--kelly-path", str(kelly_path),
        ],
        capture_output=True, text=True,
    )
    assert result.returncode in (0, 2), result.stderr
    # Top-level comparison artefacts always exist.
    assert (out_dir / "linz_frequency_source_comparison.csv").is_file()
    assert (out_dir / "linz_mwf_correlation_by_source.csv").is_file()
    mwf_corr = pd.read_csv(out_dir / "linz_mwf_correlation_by_source.csv")
    assert set(mwf_corr["frequency_source"]) == {"wordfreq", "kelly"}
    # Per-source correlation tables always exist (even if regression aborted).
    assert (out_dir / "wordfreq" / "linz_correlation_table.csv").is_file()
    assert (out_dir / "kelly" / "linz_correlation_table.csv").is_file()
